from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from typing import Any

from schemas import EXPORT_COLUMNS, QuestionItem


def load_json_file(path: Path) -> list[dict[str, Any]]:
    data = json.loads(path.read_text(encoding="utf-8-sig"))
    if isinstance(data, list):
        return [item for item in data if isinstance(item, dict)]
    if isinstance(data, dict):
        return [data]
    raise ValueError(f"Unsupported JSON payload in {path}")


def load_jsonl_file(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for line_number, line in enumerate(path.read_text(encoding="utf-8-sig").splitlines(), start=1):
        stripped = line.strip()
        if not stripped:
            continue
        item = json.loads(stripped)
        if not isinstance(item, dict):
            raise ValueError(f"{path}:{line_number} is not a JSON object")
        rows.append(item)
    return rows


def load_questions(input_dir: Path) -> list[QuestionItem]:
    if not input_dir.exists():
        raise FileNotFoundError(f"Input directory not found: {input_dir}")

    payloads: list[dict[str, Any]] = []
    for path in sorted(input_dir.glob("*.json")):
        payloads.extend(load_json_file(path))
    for path in sorted(input_dir.glob("*.jsonl")):
        payloads.extend(load_jsonl_file(path))

    questions: list[QuestionItem] = []
    validation_errors: list[str] = []
    for index, payload in enumerate(payloads, start=1):
        try:
            question = QuestionItem.from_dict(payload)
            errors = question.validate()
            if errors:
                validation_errors.append(f"item {index} ({question.question_id}): {'; '.join(errors)}")
                continue
            questions.append(question)
        except Exception as exc:
            validation_errors.append(f"item {index}: {exc}")

    if validation_errors:
        preview = "\n".join(validation_errors[:20])
        raise ValueError(f"Question validation failed:\n{preview}")

    return questions


def export_jsonl(questions: list[QuestionItem], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    lines = [json.dumps(question.to_dict(), ensure_ascii=False) for question in questions]
    output_path.write_text("\n".join(lines) + ("\n" if lines else ""), encoding="utf-8")


def export_csv(questions: list[QuestionItem], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(EXPORT_COLUMNS))
        writer.writeheader()
        for question in questions:
            writer.writerow(question.to_export_row())


def export_xlsx(questions: list[QuestionItem], output_path: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for xlsx export. Install with: pip install openpyxl") from exc

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "item_bank"
    sheet.append(list(EXPORT_COLUMNS))

    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for question in questions:
        row = question.to_export_row()
        sheet.append([row[column] for column in EXPORT_COLUMNS])

    width_by_column = {
        "A": 18,
        "B": 12,
        "C": 34,
        "D": 28,
        "E": 24,
        "F": 22,
        "G": 30,
        "H": 24,
        "I": 22,
        "J": 70,
        "K": 38,
        "L": 38,
        "M": 38,
        "N": 38,
        "O": 38,
        "P": 16,
        "Q": 70,
        "R": 45,
        "S": 45,
        "T": 45,
        "U": 45,
        "V": 45,
        "W": 45,
        "X": 15,
        "Y": 18,
        "Z": 35,
        "AA": 22,
        "AB": 22,
    }
    for column_index, _ in enumerate(EXPORT_COLUMNS, start=1):
        letter = get_column_letter(column_index)
        sheet.column_dimensions[letter].width = width_by_column.get(letter, 18)
    sheet.freeze_panes = "A2"
    workbook.save(output_path)


def export_questions(questions: list[QuestionItem], output_dir: Path, export_format: str) -> list[Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    written: list[Path] = []
    formats = ("jsonl", "csv", "xlsx") if export_format == "all" else (export_format,)

    for fmt in formats:
        if fmt == "jsonl":
            path = output_dir / "item_bank.jsonl"
            export_jsonl(questions, path)
        elif fmt == "csv":
            path = output_dir / "item_bank.csv"
            export_csv(questions, path)
        elif fmt == "xlsx":
            path = output_dir / "item_bank.xlsx"
            export_xlsx(questions, path)
        else:
            raise ValueError(f"Unsupported format: {fmt}")
        written.append(path)
    return written


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export reviewed Brenner question bank.")
    parser.add_argument("--input-dir", type=Path, default=Path("data/final_item_bank"))
    parser.add_argument("--output-dir", type=Path, default=Path("outputs"))
    parser.add_argument("--format", choices=["jsonl", "csv", "xlsx", "all"], default="all")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    questions = load_questions(args.input_dir)
    written = export_questions(questions, args.output_dir, args.format)
    for path in written:
        print(f"wrote {path}")
    print(f"exported {len(questions)} questions")


if __name__ == "__main__":
    main()
